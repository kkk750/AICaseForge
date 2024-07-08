import json
from io import BytesIO

from openpyxl.reader.excel import load_workbook


def excel2json(excel_stream):
    workbook = load_workbook(filename=BytesIO(excel_stream))
    sheets_json_data = {}

    # 遍历所有工作表
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]

        rows_data = []
        columns = []

        for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)):
            # if cell is not None:
            columns.append(cell)
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            row_data = {}
            for key, value in zip(columns, row):
                row_data[key] = value if value is not None else 'None'
            if all(value is 'None' for value in row_data.values()):
                continue
            rows_data.append(row_data)
        # 将所有值转为字符串
        rows_data = [{k: str(v) if v is not None else None for k, v in row.items()} for row in rows_data]
        sheets_json_data[sheet_name] = json.dumps(rows_data, ensure_ascii=False)
        json_data = json.dumps(sheets_json_data, ensure_ascii=False)
    return json_data


if __name__ == '__main__':
    with open('D:\\download\\测试任务划分.xlsx', 'rb') as file:
        file_stream = file.read()
    json = excel2json(file_stream)
    print(json)
